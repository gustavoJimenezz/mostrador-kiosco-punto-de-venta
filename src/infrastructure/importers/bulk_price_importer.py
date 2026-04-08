"""Adaptador de importación masiva de listas de precios de proveedores.

Lee archivos CSV o Excel (xlsx/xls) usando la stdlib ``csv`` y ``openpyxl``
(puro Python, sin instrucciones SIMD) para compatibilidad con hardware antiguo
como el Intel Core 2 Duo E5700.

Columnas requeridas:
    barcode       — Código de barras EAN-13.
    name          — Nombre/descripción del producto.
    cost_price    — Costo de compra (ej: "1.250,50" o "1250.50").

Columnas opcionales (con defaults):
    margin_percent — Margen de ganancia % (default: "30.00").
    stock          — Stock inicial (default: "0").
    min_stock      — Stock mínimo para alerta (default: "0").

Uso::

    result = BulkPriceImporter().parse("/ruta/lista.csv")
    # result.valid_rows  → list[ProductImportRow]
    # result.errors      → list[ImportRowError]
"""

from __future__ import annotations

import csv
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterator, Union

from src.application.use_cases.update_bulk_prices import ImportRowError, ProductImportRow

_REQUIRED_COLUMNS = {"barcode", "name", "cost_price"}
_UNASSIGNED = "(sin asignar)"
_IGNORE = "(ignorar)"

_OPTIONAL_DEFAULTS: dict[str, str] = {
    "margin_percent": "30",
    "stock": "0",
    "min_stock": "0",
}


@dataclass
class ImportSheet:
    """Tabla de datos en memoria (reemplaza pl.DataFrame).

    Almacena todas las celdas como strings para que ``_parse_decimal`` y
    ``_parse_int`` puedan normalizarlas independientemente del tipo original
    de la celda (número, texto, vacío).

    Attributes:
        columns: Lista de nombres de columnas en orden.
    """

    columns: list[str]
    _rows: list[dict[str, str]] = field(default_factory=list)

    def iter_rows(self, named: bool = True) -> Iterator[dict[str, str]]:
        """Itera sobre las filas como dicts ``{columna: valor_str}``.

        Args:
            named: Ignorado; existe por compatibilidad con el contrato
                   original de ``pl.DataFrame.iter_rows(named=True)``.

        Yields:
            dict con cada fila como pares columna→string.
        """
        return iter(self._rows)

    def rename(self, mapping: dict[str, str]) -> "ImportSheet":
        """Retorna un nuevo ImportSheet con columnas renombradas.

        Args:
            mapping: ``{nombre_original: nombre_nuevo}``.

        Returns:
            Nuevo ImportSheet con columnas y filas renombradas.
        """
        new_columns = [mapping.get(c, c) for c in self.columns]
        new_rows = [
            {mapping.get(k, k): v for k, v in row.items()}
            for row in self._rows
        ]
        return ImportSheet(new_columns, new_rows)

    @classmethod
    def from_dict(cls, data: dict[str, list]) -> "ImportSheet":
        """Crea un ImportSheet desde un dict de listas (útil en tests).

        Args:
            data: ``{nombre_columna: [valor1, valor2, ...]}``.
                  Los valores se convierten a string; None → "".

        Returns:
            ImportSheet con una fila por elemento de las listas.
        """
        columns = list(data.keys())
        if not columns:
            return cls([], [])
        n = len(next(iter(data.values())))
        rows = [
            {
                col: str(data[col][i]) if data[col][i] is not None else ""
                for col in columns
            }
            for i in range(n)
        ]
        return cls(columns, rows)


# ---------------------------------------------------------------------------
# Helpers de lectura de archivos (puro Python, sin SIMD)
# ---------------------------------------------------------------------------


def _cell_to_str(value: object) -> str:
    """Convierte un valor de celda (openpyxl/xlrd) a string.

    Para floats enteros (ej: 1250.0) retorna "1250" en lugar de "1250.0",
    para no confundir al parser de formato argentino.

    Args:
        value: Valor de celda de cualquier tipo Python.

    Returns:
        Representación string del valor; "" si es None.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)


def _read_csv(path: Path) -> ImportSheet:
    """Lee un archivo CSV como ImportSheet usando la stdlib.

    Usa ``utf-8-sig`` para tolerar el BOM que Excel agrega al exportar CSV UTF-8.

    Args:
        path: Ruta al archivo .csv.

    Returns:
        ImportSheet con todas las celdas como strings.
    """
    with open(path, encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if reader.fieldnames is None:
            return ImportSheet([], [])
        columns = list(reader.fieldnames)
        rows = [{k: (v or "") for k, v in row.items()} for row in reader]
    return ImportSheet(columns, rows)


def _read_xlsx(path: Path) -> ImportSheet:
    """Lee un archivo .xlsx como ImportSheet usando openpyxl (puro Python).

    Args:
        path: Ruta al archivo .xlsx.

    Returns:
        ImportSheet con todas las celdas como strings.
    """
    from openpyxl import load_workbook  # importación diferida

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        return ImportSheet([], [])

    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return ImportSheet([], [])

    columns = [
        str(cell) if cell is not None else f"col_{i}"
        for i, cell in enumerate(all_rows[0])
    ]

    data_rows: list[dict[str, str]] = []
    for row in all_rows[1:]:
        row_dict: dict[str, str] = {}
        for i, cell in enumerate(row):
            key = columns[i] if i < len(columns) else f"col_{i}"
            row_dict[key] = _cell_to_str(cell)
        data_rows.append(row_dict)

    return ImportSheet(columns, data_rows)


def _read_xls(path: Path) -> ImportSheet:
    """Lee un archivo .xls (Excel 97-2003) como ImportSheet usando xlrd.

    ``xlrd`` es una dependencia opcional. Si no está instalado se lanza
    ``ImportError`` con instrucciones de instalación.

    Args:
        path: Ruta al archivo .xls.

    Returns:
        ImportSheet con todas las celdas como strings.

    Raises:
        ImportError: Si ``xlrd`` no está instalado.
    """
    try:
        import xlrd
    except ImportError as exc:
        raise ImportError(
            "Para leer archivos .xls (Excel 97-2003) instalar: "
            "poetry add 'xlrd>=1.2,<2.0'"
        ) from exc

    wb = xlrd.open_workbook(str(path))
    ws = wb.sheet_by_index(0)

    if ws.nrows == 0:
        return ImportSheet([], [])

    columns = [str(ws.cell_value(0, col)) for col in range(ws.ncols)]

    data_rows: list[dict[str, str]] = []
    for row_idx in range(1, ws.nrows):
        row_dict = {
            columns[col]: _cell_to_str(ws.cell_value(row_idx, col))
            for col in range(ws.ncols)
        }
        data_rows.append(row_dict)

    return ImportSheet(columns, data_rows)


# ---------------------------------------------------------------------------
# DTOs
# ---------------------------------------------------------------------------


@dataclass
class ParseResult:
    """Resultado del parsing de un archivo CSV/Excel.

    Attributes:
        valid_rows: Filas correctamente validadas, listas para el use case.
        errors: Errores por fila (no abortan el lote).
    """

    valid_rows: list[ProductImportRow] = field(default_factory=list)
    errors: list[ImportRowError] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Importer
# ---------------------------------------------------------------------------


class BulkPriceImporter:
    """Adaptador de importación de listas de precios de proveedores.

    Lee CSV o Excel, valida schema y filas, y produce DTOs limpios.
    Los errores de fila se acumulan sin abortar el lote completo.

    Implementado con stdlib ``csv`` y ``openpyxl`` (puro Python) para
    compatibilidad con CPUs que no soporten instrucciones SIMD modernas
    (AVX2, SSE4.2) como el Intel Core 2 Duo E5700.
    """

    def load(self, file_path: Union[str, Path]) -> ImportSheet:
        """Lee un archivo CSV o Excel y retorna un ImportSheet sin validar.

        Útil para obtener columnas y filas de preview antes de importar.

        Args:
            file_path: Ruta al archivo CSV (.csv) o Excel (.xlsx/.xls).

        Returns:
            ImportSheet con columnas y filas como strings.

        Raises:
            ValueError: Si el archivo no tiene extensión soportada.
            FileNotFoundError: Si el archivo no existe.
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {path}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            return _read_csv(path)
        elif suffix == ".xlsx":
            return _read_xlsx(path)
        elif suffix == ".xls":
            return _read_xls(path)
        else:
            raise ValueError(
                f"Extensión no soportada: '{suffix}'. Use .csv, .xlsx o .xls."
            )

    def parse(self, file_path: Union[str, Path]) -> ParseResult:
        """Lee y valida un archivo CSV o Excel de lista de precios.

        Args:
            file_path: Ruta al archivo CSV (.csv) o Excel (.xlsx/.xls).

        Returns:
            ParseResult con filas válidas y errores por fila.

        Raises:
            ValueError: Si el archivo no tiene extensión soportada.
            FileNotFoundError: Si el archivo no existe.
            Exception: Si el archivo tiene formato inválido.
        """
        path = Path(file_path)

        if not path.exists():
            raise FileNotFoundError(f"Archivo no encontrado: {path}")

        suffix = path.suffix.lower()
        if suffix == ".csv":
            sheet = _read_csv(path)
        elif suffix == ".xlsx":
            sheet = _read_xlsx(path)
        elif suffix == ".xls":
            sheet = _read_xls(path)
        else:
            raise ValueError(
                f"Extensión no soportada: '{suffix}'. Use .csv, .xlsx o .xls."
            )

        return self.parse_dataframe(sheet)

    def parse_dataframe(
        self,
        df: ImportSheet,
        column_mapping: dict[str, str] | None = None,
        global_margin: Decimal | None = None,
    ) -> ParseResult:
        """Valida y construye DTOs desde un ImportSheet ya cargado.

        Si ``column_mapping`` se provee (``{campo_destino: col_archivo}``),
        renombra las columnas antes de validar. Las entradas cuya columna sea
        ``_UNASSIGNED`` o ``_IGNORE`` se omiten. El campo destino ``net_cost``
        se traduce internamente a ``cost_price`` para el validador.

        Si ``global_margin`` se provee, sobreescribe incondicionalmente el
        ``margin_percent`` de cada fila, ignorando el valor del archivo.

        Args:
            df: ImportSheet con todos los datos como strings.
            column_mapping: Mapeo ``{campo_destino: nombre_columna_archivo}``.
                            Campos destino: ``barcode``, ``name``, ``net_cost``,
                            ``category``.
                            Si es None, se usa el ImportSheet sin modificaciones.
            global_margin: Margen de ganancia porcentual a aplicar a todas las
                           filas. Si es None, se usa el valor del archivo o el
                           default de 30%.

        Returns:
            ParseResult con filas válidas y errores acumulados.
        """
        if column_mapping:
            rename_map: dict[str, str] = {}
            _ALIAS = {"net_cost": "cost_price"}

            for dest_field, file_col in column_mapping.items():
                if file_col in {_UNASSIGNED, _IGNORE, ""}:
                    continue
                if file_col not in df.columns:
                    continue
                internal = _ALIAS.get(dest_field, dest_field)
                rename_map[file_col] = internal

            if rename_map:
                df = df.rename(rename_map)

        return self._validate_and_build(df, global_margin=global_margin)

    def _validate_and_build(
        self,
        df: ImportSheet,
        global_margin: Decimal | None = None,
    ) -> ParseResult:
        """Valida el schema del ImportSheet y construye los DTOs.

        Args:
            df: ImportSheet con todos los datos como strings.
            global_margin: Si se provee, sobreescribe el margen de cada fila.

        Returns:
            ParseResult con filas válidas y errores acumulados.
        """
        result = ParseResult()

        missing = _REQUIRED_COLUMNS - set(df.columns)
        if missing:
            raise ValueError(
                f"Columnas requeridas ausentes: {', '.join(sorted(missing))}. "
                f"Columnas presentes: {', '.join(df.columns)}"
            )

        for row_idx, row_data in enumerate(df.iter_rows(named=True), start=2):
            self._process_row(row_idx, row_data, result, global_margin=global_margin)

        return result

    def _process_row(
        self,
        row_number: int,
        row_data: dict,
        result: ParseResult,
        global_margin: Decimal | None = None,
    ) -> None:
        """Valida y convierte una fila individual.

        Errores de validación se agregan a ``result.errors`` sin lanzar excepción.

        Args:
            row_number: Número de fila en el archivo original (base 2 por encabezado).
            row_data: Diccionario con los valores de la fila como strings.
            result: ParseResult donde acumular válidos o errores.
            global_margin: Si se provee, reemplaza incondicionalmente el margen de la fila.
        """
        barcode = (row_data.get("barcode") or "").strip()

        if not barcode:
            result.errors.append(
                ImportRowError(
                    row_number=row_number,
                    barcode="",
                    reason="Campo 'barcode' vacío o ausente.",
                )
            )
            return

        name = (row_data.get("name") or "").strip()
        if not name:
            result.errors.append(
                ImportRowError(
                    row_number=row_number,
                    barcode=barcode,
                    reason="Campo 'name' vacío o ausente.",
                )
            )
            return

        cost_raw = (row_data.get("cost_price") or "").strip()
        cost_price = self._parse_decimal(cost_raw)
        if cost_price is None or cost_price <= Decimal("0"):
            result.errors.append(
                ImportRowError(
                    row_number=row_number,
                    barcode=barcode,
                    reason=f"'cost_price' inválido o <= 0: '{cost_raw}'.",
                )
            )
            return

        if global_margin is not None:
            margin_percent = global_margin
        else:
            margin_raw = (
                row_data.get("margin_percent") or _OPTIONAL_DEFAULTS["margin_percent"]
            ).strip()
            margin_percent = self._parse_decimal(margin_raw)
            if margin_percent is None or margin_percent < Decimal("0"):
                result.errors.append(
                    ImportRowError(
                        row_number=row_number,
                        barcode=barcode,
                        reason=f"'margin_percent' inválido: '{margin_raw}'.",
                    )
                )
                return

        stock = self._parse_int(
            (row_data.get("stock") or _OPTIONAL_DEFAULTS["stock"]).strip()
        )
        if stock is None:
            result.errors.append(
                ImportRowError(
                    row_number=row_number,
                    barcode=barcode,
                    reason=f"'stock' inválido: '{row_data.get('stock')}'.",
                )
            )
            return

        min_stock = self._parse_int(
            (row_data.get("min_stock") or _OPTIONAL_DEFAULTS["min_stock"]).strip()
        )
        if min_stock is None:
            result.errors.append(
                ImportRowError(
                    row_number=row_number,
                    barcode=barcode,
                    reason=f"'min_stock' inválido: '{row_data.get('min_stock')}'.",
                )
            )
            return

        category_name = (row_data.get("category") or "").strip()

        result.valid_rows.append(
            ProductImportRow(
                barcode=barcode,
                name=name,
                cost_price=cost_price,
                margin_percent=margin_percent,
                stock=stock,
                min_stock=min_stock,
                source_row=row_number,
                category_name=category_name,
            )
        )

    @staticmethod
    def _parse_decimal(value: str) -> Decimal | None:
        """Convierte un string numérico a Decimal detectando el formato automáticamente.

        Heurística:
            - Si contiene coma → formato argentino (coma=decimal, punto=miles).
              Ej: "1.250,50" → Decimal("1250.50"), "843,00" → Decimal("843.00").
            - Si no contiene coma → formato inglés o entero (punto=decimal).
              Ej: "843.00" → Decimal("843.00"), "1250" → Decimal("1250").

        Args:
            value: String numérico a convertir.

        Returns:
            Decimal si la conversión es exitosa, None si el formato es inválido.
        """
        if not value:
            return None
        try:
            if "," in value:
                normalized = value.replace(".", "").replace(",", ".")
            else:
                normalized = value
            return Decimal(normalized)
        except InvalidOperation:
            return None

    @staticmethod
    def _parse_int(value: str) -> int | None:
        """Convierte un string a entero.

        Args:
            value: String numérico entero.

        Returns:
            int si la conversión es exitosa, None si el formato es inválido.
        """
        if not value:
            return 0
        try:
            return int(value)
        except ValueError:
            return None
